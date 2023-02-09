import asyncio
import datetime
import hashlib
import json
import logging
import os
import re
import time
import typing

from openpyxl import load_workbook

from .config import ADMINS
from .utils import safedate, safestr

logger = logging.getLogger(__name__)


class Notifier:
    def __init__(self):
        self._db_path: str = os.path.join(os.path.dirname(__file__), "notifier.json")
        self._db: dict = {}
        self._queue: list = []
        self._bot: "Bot" = None  # type: ignore # noqa: F821

        self._load()

    def set_bot(self, bot: "Bot"):  # type: ignore # noqa: F821
        self._bot = bot

    def _load(self):
        if os.path.exists(self._db_path):
            with open(self._db_path, "r") as f:
                self._db = json.load(f)

        self._db.setdefault("fluorography", {})
        self._db.setdefault("mantoux", {})
        self._db.setdefault("DPT", {})

    def _save(self):
        with open(self._db_path, "w") as f:
            json.dump(self._db, f)

    def _remind(
        self,
        type_: str,
        name: str,
        age: int,
        grade: str,
        last_date: datetime.datetime,
    ):
        logger.info("Reminding %s about %s", name, type_)

        type_formatted = {
            "fluorography": "флюорографию",
            "mantoux": "диаскинтест",
            "DPT": "АДС-М",
        }[type_]

        type_emoji = {
            "fluorography": "🩻",
            "mantoux": "📏",
            "DPT": "💉",
        }[type_]

        last_date_formatted = (
            f" (последний раз было <b>{last_date.strftime('%d.%m.%Y')}</b>)"
            if isinstance(last_date, datetime.datetime)
            else " (последний раз было <i>никогда</i>)"
        )

        grade = grade.upper()
        grade_formatted = f"{grade[:-1]} «{grade[-1]}»"

        self._queue += [
            (
                f"{type_emoji} <b>{name}</b> {grade_formatted} ({age} лет) необходимо"
                f" дать направление на <b>{type_formatted}</b>{last_date_formatted}"
            ),
        ]

    async def _flush(self, force_send: typing.Optional[int] = None):
        if not self._queue:
            return

        while self._queue:
            chunk = ""
            while self._queue and len(chunk) + len(self._queue[0]) < 4096:
                chunk += self._queue.pop(0) + "\n➖➖➖➖➖➖➖➖➖➖\n"

            if force_send:
                await self._bot.send_message(force_send, chunk)
            else:
                for admin in ADMINS:
                    try:
                        await self._bot.send_message(admin, chunk)
                    except Exception:
                        logger.exception("Can't send")

    async def poll(self, force_send: int = False):
        while True:
            try:
                wb = load_workbook(
                    os.path.join(os.path.dirname(__file__), "docx", "database.xlsx")
                )
                ws = wb.active
                for row in ws.iter_rows(min_row=2):
                    if row[0].value is None:
                        continue

                    name = " ".join(
                        (
                            safestr(row[1].value),
                            safestr(row[2].value),
                            safestr(row[3].value),
                        )
                    )
                    user_id: str = hashlib.md5(name.lower().encode()).hexdigest()
                    date_of_birth: datetime.datetime = safedate(row[5].value)
                    age: int = (datetime.datetime.now() - date_of_birth).days // 365

                    if (
                        age >= 15
                        and (
                            not safedate(row[10].value)
                            or (datetime.datetime.now() - safedate(row[10].value)).days
                            > 365
                        )
                        and (
                            force_send
                            or user_id not in self._db["fluorography"]
                            or (time.time() - self._db["fluorography"][user_id])
                            > 60 * 60 * 24 * 30
                        )
                    ):
                        self._remind(
                            "fluorography",
                            name,
                            age,
                            safestr(row[0].value),
                            safedate(row[10].value),
                        )
                        self._db["fluorography"][user_id] = round(time.time())

                    if (
                        age >= 14
                        and (
                            not safedate(row[7].value)
                            or safedate(row[7].value).year < date_of_birth.year + 14
                        )
                        and (
                            force_send
                            or user_id not in self._db["DPT"]
                            or (time.time() - self._db["DPT"][user_id])
                            > 60 * 60 * 24 * 30
                        )
                    ):
                        self._remind(
                            "DPT",
                            name,
                            age,
                            safestr(row[0].value),
                            safedate(row[7].value),
                        )
                        self._db["DPT"][user_id] = round(time.time())

                    try:
                        last_mantoux = re.findall(
                            r"\d{1,2}\.\d{1,2}\.\d{2,4}",
                            safestr(row[6].value),
                        )[-1]
                    except IndexError:
                        last_mantoux = "01.01.1970"

                    day, month, year = last_mantoux.split(".")
                    if len(year) == 2:
                        year = "20" + year

                    if len(day) == 1:
                        day = "0" + day

                    if len(month) == 1:
                        month = "0" + month

                    last_mantoux = f"{day}.{month}.{year}"

                    last_mantoux = datetime.datetime.strptime(last_mantoux, "%d.%m.%Y")
                    if (
                        age < 15
                        and (datetime.datetime.now() - last_mantoux).days > 365
                        and (
                            force_send
                            or user_id not in self._db["mantoux"]
                            or (time.time() - self._db["mantoux"][user_id])
                            > 60 * 60 * 24 * 30
                        )
                    ):
                        self._remind(
                            "mantoux",
                            name,
                            age,
                            safestr(row[0].value),
                            last_mantoux,
                        )
                        self._db["mantoux"][user_id] = round(time.time())

                await self._flush(force_send)
                self._save()
            except Exception:
                logger.exception("Failed to poll")

            if force_send:
                break

            await asyncio.sleep(60)
